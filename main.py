import openpyxl, re
import openpyxl.styles
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import Select
from openpyxl.styles import Alignment, Font
from time import sleep
from bs4 import BeautifulSoup

def wait_url(driver: webdriver.Chrome, url: str):
    while True:
        cur_url = driver.current_url
        if cur_url == url:
            break
        sleep(0.1)

def find_element(driver: webdriver.Chrome, whichBy, unique: str) -> WebElement:
    while True:
        try:
            element = driver.find_element(whichBy, unique)
            break
        except:
            pass
        sleep(1)
    return element

def find_elements(driver : webdriver.Chrome, whichBy, unique: str) -> list[WebElement]:
    while True:
        try:
            elements =driver.find_elements(whichBy, unique)
            break
        except:
            pass
        sleep(1)
    return elements

remote_debugging_address = "127.0.0.1:9024"
chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option("debuggerAddress", remote_debugging_address)
driver = webdriver.Chrome(options=chrome_options)

url ="https://app.atplquestions.com/test"

email = "kakou324@gmail.com"
password = "Upwork123@"
match_num = 0


# driver.get(url)
wait_url(driver, url)
sleep(1)



questions_num = find_element(driver, By.CLASS_NAME, "maxSize")
questions_num_int = int(questions_num.text)
print(questions_num_int)

start_study = find_element(driver, By.CLASS_NAME, "startStudy")
start_study.click()
print("Started!")


file_name = find_element(driver, By.CLASS_NAME, "testname").text
print(file_name)
parts = file_name.split('|')
# Extract the second part (index 1) and split it by space
code_parts = parts[1].strip().split()
# The code is the first part of the second split
code = code_parts[0]
print(code)

wb = openpyxl.Workbook()
ws = wb.active

ws['A1'] = "Question number"
ws['A1'].font = Font(bold=True)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws['B1'] = "Question"
ws['B1'].font = Font(bold=True)
ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
ws['C1'] = "option1"
ws['C1'].font = Font(bold=True)
ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
ws['D1'] = "option2"
ws['D1'].font = Font(bold=True)
ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
ws['E1'] = "option3"
ws['E1'].font = Font(bold=True)
ws['E1'].alignment = Alignment(horizontal='center', vertical='center')
ws['F1'] = "option4"
ws['F1'].font = Font(bold=True)
ws['F1'].alignment = Alignment(horizontal='center', vertical='center')
ws['G1'] = "answer"
ws['G1'].font = Font(bold=True)
ws['G1'].alignment = Alignment(horizontal='center', vertical='center')
ws['H1'] = "note"
ws['H1'].font = Font(bold=True)
ws['H1'].alignment = Alignment(horizontal='center', vertical='center')

wb.save(f'{code}.xlsx')

for _ in range(questions_num_int):
    match_num +=1
    workbook = openpyxl.load_workbook(f'{code}.xlsx')
    sheet = workbook['Sheet']
    sleep(1)
    ques_mum =find_element(driver, By.CLASS_NAME, "q-number").text
    sheet[f'A{match_num+1}'] = ques_mum
    quesiton = find_element(driver, By.CLASS_NAME, "questionText").text
    sheet[f'B{match_num+1}'] = quesiton
    quesiton_options = find_elements(driver, By.CLASS_NAME, "questionOption")
    for index, question_each_option in enumerate(quesiton_options):
        question_option = question_each_option.find_element(By.CLASS_NAME, "texter")
        if index == 0:
            sheet[f'C{match_num+1}'] = question_option.text
        elif index == 1:
            sheet[f'D{match_num+1}'] = question_option.text
        elif index == 2:
            sheet[f'E{match_num+1}'] = question_option.text
        elif index == 3:
            sheet[f'F{match_num+1}'] = question_option.text

    user_navs = find_elements(driver, By.CLASS_NAME, "nav-link")
    for explanation in user_navs:
        if explanation.text == "Explanation":
            explanation.click()
            sleep(2)
    explanation_txt = find_element(driver, By.CLASS_NAME, "exp-text").get_attribute('innerHTML')
    soup = BeautifulSoup(explanation_txt, 'html.parser')
    # plain_text = ""
    # for element in soup.find_all():
    #     if element.name == 'strong':
    #         plain_text += fontstyle.apply(element.get_text(), 'bold')
    #     elif element.name == 'u':
    #         plain_text += fontstyle.apply(element.get_text(), 'underline')
    #     elif element.name == 'span' and 'letter-spacing' in element.get('style', ''):
    #         letter_spacing = element['style'].split('letter-spacing:')[1].split(';')[0]
    #         plain_text += fontstyle.apply(element.get_text(), f'letter-spacing:{letter_spacing}')
    #     else:
    #         plain_text += element.get_text()
    #     plain_text += '\n'
    plain_text = ""
    for element in soup.recursiveChildGenerator():
        if isinstance(element, str):
            plain_text += element
        elif element.name == 'strong':
            plain_text += '<bold>' + element.text + '</bold>'
        elif element.name == 'u':
            plain_text += '<underline>' + element.text + '</underline>'
        # elif element.name == 'em':
        #     plain_text += '<italic>' + element.text + '</italic>'
    cleand_text = re.sub(r'\x1b\[[0-9;]*m', '', plain_text)

    sheet[f'H{match_num+1}'] = cleand_text
    # sheet[f'H{match_num+1}'] = soup.get_text()
    # for element in soup.find_all():
    #     text = element.get_text()
    #     if element.name == 'span':
    #         if 'font-weight' in element.get('style', ''):
    #             sheet[f'H{match_num+1}'].font = Font(bold=True)
    #         if 'color' in element.get('style', ''):
    #             sheet[f'H{match_num+1}'].font = Font(color=element['style'].split('color:')[1].split(';')[0])
    #     # sheet[f'H{match_num+1}'].value = text
    find_element(driver, By.CLASS_NAME, "next").click()
    workbook.save(f'{code}.xlsx')
